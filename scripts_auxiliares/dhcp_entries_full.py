import openpyxl
import xlrd
import csv
import unicodedata
import json
import ipaddress
import netaddr

def nextMAC(mac):
	mac=netaddr.EUI(mac)
	next_mac=netaddr.EUI(int(mac)+1)
	next_mac.dialect=netaddr.mac_unix_expanded
	return str(next_mac).upper()

def getData(db,sheet):
	wb=openpyxl.load_workbook(db)
	data=wb[sheet]
	return data

def mapColNum(col):
	col=col.lower()
	letters=list(col)
	if len(letters)>1:
		col_num=((int(ord(letters[0]))-96)*26)+(int(ord(letters[1]))-96)
		return col_num
	else:
		col_num=int(ord(letters[0]))-96
		return col_num

def formatMACaddress(mac):
	mac=str(mac)
	format_mac=""
	if ':' in mac:
		return mac
	else:
		for i in range(len(mac)):
			if i%2==0 or i==len(mac)-1:
				format_mac=format_mac+mac[i]
			elif i%2==1:
				format_mac=format_mac+mac[i]+':'
		return format_mac

def iscDHCPformat(leases):
	all_entries=""
	for key in leases.keys():
		#print key
		subnet=ipaddress.IPv4Network(key)
		subnet_address=subnet.network_address
		netmask=subnet.netmask
		gateway=list(subnet.hosts())[0]
		broadcast=subnet.broadcast_address
		entry_data=[subnet_address,netmask,gateway,broadcast]
		hosts=[]
		for i in range(len(leases[key])):
			#print leases[key][i]
			host_entry="\thost "+leases[key][i][1]+" {hardware ethernet %(mac_address)s; fixed-address %(ip_address)s;}"%{"mac_address":nextMAC(formatMACaddress(leases[key][i][1])),"ip_address":leases[key][i][0]}
			hosts.append(host_entry)
		group_entry="{"
		for j in range(len(hosts)):
			group_entry=group_entry+"\n\t"+hosts[j-1]
			if j==len(hosts)-1:
				group_entry=group_entry+"\n\t}\n"
		entry_string="subnet %(subnet)s netmask %(netmask)s { \n\toption routers %(gateway)s;\n\toption broadcast-address %(broadcast)s;\n\toption domain-name-servers 189.247.96.1, 189.247.97.1;\n\toption subnet-mask %(netmask)s;\n\tgroup %(group_entry)s"%{"subnet":subnet_address,"netmask":netmask,"gateway":gateway,"broadcast":broadcast,"group_entry":group_entry}
		entry_string=entry_string+"}\n"
		all_entries=all_entries+entry_string
	return all_entries

def getNetworks(db,sheet):
	wb=openpyxl.load_workbook(db)
	data=wb[sheet]
	i=2
	networks={}
	while i<data.max_row:
		#print i
		#print data.max_row
		current_site=data.cell(i,mapColNum('A')).value
		next_site=data.cell(i+1,mapColNum('A')).value
		if current_site!=next_site:
			networks[current_site]=data.cell(i,mapColNum('B')).value
		i+=1
	return networks

networks=getNetworks('ref_net.xlsx','Sheet1')

print 'all networks loaded'

data=getData('MAC.xlsx','Sheet1')
i=2
all_devices={}
while i<data.max_row:
	site_devices=[]
	current_inegi=data.cell(i,mapColNum('D')).value
	next_inegi=data.cell(i+1,mapColNum('D')).value
	current_device=data.cell(i,mapColNum('B')).value
	site_devices.append(current_device)
	while current_inegi==next_inegi:
		i+=1
		current_inegi=data.cell(i,mapColNum('D')).value
		next_inegi=data.cell(i+1,mapColNum('D')).value
		current_device=data.cell(i,mapColNum('B')).value
		site_devices.append(current_device)
	all_devices[current_inegi]=site_devices
	i+=1

for key in all_devices.keys():
	print key
	print all_devices[key]

print 'all devices loaded'

all_leases={}
for key in all_devices.keys():
	site_network=list(v for k,v in networks.iteritems() if str(key) in k)
	site_leases=[]
	if len(site_network)==1:
		ip_network=ipaddress.IPv4Network(site_network[0]+'/28')
		site_devices=all_devices[key]
		i=0
		while i<len(site_devices):
			lease=ip_network[i+2],site_devices[i]
			site_leases.append(lease)
			i+=1
		all_leases.update({ip_network:site_leases})

print all_leases
print iscDHCPformat(all_leases)



#for key in all_leases.keys():
#	print key
#	print all_leases[key]

#for key in all_devices.keys():
#	print key
#	print all_devices[key]

#for key in networks.keys():
#	print key
#	print networks[key]
#print 'networks loaded'
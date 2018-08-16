import openpyxl
import xlrd
import csv
import unicodedata
import json

def mapColNum(col):
	col=col.lower()
	letters=list(col)
	if len(letters)>1:
		col_num=((int(ord(letters[0]))-96)*26)+(int(ord(letters[1]))-96)
		return col_num
	else:
		col_num=int(ord(letters[0]))-96
		return col_num

def getData(db):
	wb=openpyxl.load_workbook(db)
	data=wb['Sheet1']
	all_config_list={}
	i=4
	while i<=data.max_row:
		inegi_lines_pre=[]
		inegi_lines_pos=[]
		inegi_config={}
		if isinstance(data.cell(i,mapColNum('B')).value,unicode) and str(data.cell(i,mapColNum('F')).value).lower()=='agencia':
			current_cve_inegi=data.cell(i,mapColNum('B')).value
			#print current_cve_inegi
			next_cve_inegi=data.cell(i+1,mapColNum('B')).value
			current_line=data.cell(i,mapColNum('C')).value
			current_password=data.cell(i,mapColNum('D')).value
			line_mode=str(data.cell(i,mapColNum('J')).value)
			if line_mode=='PREPAGO':
				inegi_lines_pre.append((current_line,current_password))
				print str(current_line)+' prepago'
			elif line_mode=='POSPAGO':
				inegi_lines_pos.append((current_line,current_password))
				print str(current_line)+' pospago'
			else:
				inegi_lines_pos.append((current_line,current_password))
				print str(current_line)+' sin info'
				#indicar que no se encontro el tipo de servicio
			while current_cve_inegi==next_cve_inegi:
				i+=1
				current_cve_inegi=data.cell(i,mapColNum('B')).value
				next_cve_inegi=data.cell(i+1,mapColNum('B')).value
				current_line=data.cell(i,mapColNum('C')).value
				line_mode=str(data.cell(i,mapColNum('J')).value)
				current_password=data.cell(i,mapColNum('D')).value
				if line_mode=='PREPAGO':
					inegi_lines_pre.append((current_line,current_password))
					print str(current_line)+' prepago'
				elif line_mode=='POSPAGO':
					inegi_lines_pos.append((current_line,current_password))
					print str(current_line)+' pospago'
				else:
					inegi_lines_pos.append((current_line,current_password))
					print str(current_line)+' sin info'
			#print len(inegi_lines_pos)
			modems_pos=len(inegi_lines_pos)/2
			if len(inegi_lines_pos)%2==1:
				#print 'entro al if'
				modems_pos+=1
			modems_pre=len(inegi_lines_pre)/2
			if len(inegi_lines_pre)%2==1:
				modems_pre+=1
			#print modems_pre
			#print modems_pos
			for k in range(modems_pos):
				inegi_config['modem_pos_'+str(k)]=[]
				try:
					inegi_config['modem_pos_'+str(k)].append(inegi_lines_pos[k*2])
					inegi_config['modem_pos_'+str(k)].append(inegi_lines_pos[(k*2)+1])
				except IndexError,e:
					#se terminaron las lineas de pospago
					a=1
			for k in range(modems_pre):
				inegi_config['modem_pre_'+str(k)]=[]
				try:
					inegi_config['modem_pre_'+str(k)].append(inegi_lines_pre[k*2])
					inegi_config['modem_pre_'+str(k)].append(inegi_lines_pre[(k*2)+1])
				except IndexError,e:
					#se terminaron las lineas de prepago
					a=1
			#print 'modem_pos_'+str(k)
			try:
				#print k
				#print len(inegi_config['modem_pos_'+str(k)])
				if len(inegi_config['modem_pos_'+str(k)])<1:
					delete_key='modem_pos_'+str(k)
					#print delete_key
					try:
						if delete_key in inegi_config: del inegi_config[delete_key]
					except NameError,e:
						a=1
				#print current_cve_inegi
				#print
				#print inegi_config
				all_config_list.update({current_cve_inegi:inegi_config})
			except KeyError,e:
				a=1
			try:
				if len(inegi_config['modem_pre_'+str(k)])<1:
					delete_key='modem_pre_'+str(k)
					#print delete_key
					try:
						if delete_key in inegi_config: del inegi_config[delete_key]
					except NameError,e:
						a=1
				#print current_cve_inegi
				#print
				#print inegi_config
			except KeyError,e:
				a=1
			proxy=str(data.cell(i,mapColNum('I')).value)
			inegi_config.update({'proxy':proxy})
			all_config_list.update({current_cve_inegi:inegi_config})
			# try:
			# 	if len(inegi_config['modem_pos_'+str(k)])<1:
			# 		delete_key='modem_pos_'+str(k)
			# 		if delete_key in inegi_config: del inegi_config[delete_key]
			# 	all_config_list.update({current_cve_inegi:inegi_config})
			# except NameError,e:
			# 	print 'esta es foncos'
		#traer los parametros que faltan
		i+=1
		#print i
		#print 'hola'
	return all_config_list

all_config_list=getData('BD.xlsx')
for inegi in all_config_list.keys():
	#print all_config_list[inegi]
	for key in all_config_list[inegi].keys():
		config_dict={}
		if 'modem' in key:
			#parametros linea 1
			config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.URI':'+52'+str(all_config_list[inegi][key][0][0])})
			config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.AuthPassword':str(all_config_list[inegi][key][0][1])})
			config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.AuthUserName':'+52'+str(all_config_list[inegi][key][0][0])+'@ims.telmex.com'})
			config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.Enable':'Enabled'})
			config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.X_HUAWEI_DTMFMethod':'InBand'})
			
			#parametros linea 2
			try:
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.URI':'+52'+str(all_config_list[inegi][key][1][0])})
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.AuthPassword':str(all_config_list[inegi][key][1][1])})
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.AuthUserName':'+52'+str(all_config_list[inegi][key][1][0])+'@ims.telmex.com'})
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.Enable':'Enabled'})
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.X_HUAWEI_DTMFMethod':'InBand'})
			except IndexError:
				pass
			#parametros perfil de voz
			try:
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.SIP.OutboundProxy':all_config_list[inegi]['proxy']})
			except KeyError:
				pass
			if 'pre' in key:
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.DigitMap':'*4001*'})
			else:
				config_dict.update({'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.DigitMap':'[1-9]XXXXXXX|***XX|XX*X.#|01XXXXXXXXXX|04[4-5]XXXXXXXXXX|001XXXXXXXXXX|*86|*88|02[1-9]XXXXXXXXX|020|030|03[2-9]|031[0-2]XXX|031[3-4]XXXX|040|0[5-8]X|090|091XXXXXXXXXX|09[2-9]XX.T|00[2-9]XX.T|*#XX*XX.*XX.#|*#XX*XX.#|*#XX#|*#56*|*#26*|*[0-79]X#|*[0-79]X*XX.#|*[0-79]X*XX.*XX.#|*[0-79]X*XX.*XX.*XX.#|*56*|*66*|*69|*89|#XX#|#XX*XX.#|#XX*XX.*XX.#|#*XX*XX.*XX.#|#*XX*XX.#|**X#|163|96163|1471|1475|14713|*34*1234|[1-9]X|[1-9]XX|[1-9]XXX|[1-9]XXXX|[1-9]XXXXX|[1-9]XXXXXX|*222XXX|*9999|**XX|#X.T|##X.T|**X.T|*[0-79]X.T|*8[0-579]X.T|*8[0-579].T'})

			#parametros wifi
			config_dict.update({'InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.PreSharedKey.1.PreSharedKey':'m4n4g3m3nt'})
			config_dict.update({'InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.SSID':'Elara_mgmt'})
			config_dict.update({'InternetGatewayDevice.LANDevice.1.LANHostConfigManagement.DHCPServerEnable':False})
			config_dict.update({'InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.Enable':False})

			#parametros portal de gestion
			config_dict.update({'InternetGatewayDevice.UserInterface.X_HUAWEI_Web.UserInfo.1.Username':'ELARA'})
			config_dict.update({'InternetGatewayDevice.UserInterface.X_HUAWEI_Web.UserInfo.1.Userpassword':'El4r4com!'})

			#tr069 elara
			config_dict.update({'InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.AddressingType':'DHCP'})
			config_dict.update({'InternetGatewayDevice.ManagementServer.URL':'https://172.20.111.202'})

			with open('configs/'+inegi+key+'.txt','w') as outfile:
				json.dump(config_dict,outfile)
			print inegi+key
# for inegi in all_config_list.keys():
# 	pre_counter=0
# 	pos_counter=0
# 	csv_list=[]
# 	csv_list.append(inegi[0])
# 	device_number=0
# 	for device in all_config_list[inegi].keys():
# 		if 'pre' in device:
# 			pre_counter+=len(all_config_list[inegi][device])
# 			device_number+=1
# 		elif 'pos' in device:
# 			pos_counter+=len(all_config_list[inegi][device])
# 			device_number+=1
# 	csv_list.append(str(pre_counter))
# 	csv_list.append(str(pos_counter))
# 	csv_list.append(str(device_number))
# 	print csv_list
# 	with open('reporte.csv','a') as outfile:
# 		wr=csv.writer(outfile,quoting=csv.QUOTE_ALL,lineterminator='\n')
# 		wr.writerow(csv_list)
# 		outfile.close()	


#notas yafte
#reducir timer a 6 minutos
#corroborar conectividad con hg antes de configurar
#verificar 3 veces en lugar de 4
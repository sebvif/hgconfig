from flask import render_template, redirect, flash, url_for
from app import app
from app.forms import ConfigureForm, VerifyForm, ConfirmForm, GoBackForm
import pycurl
from datetime import datetime, timedelta
import urllib
from StringIO import StringIO
import json
import openpyxl
import xlrd
import logging
import time
from platform   import system as system_name  # Returns the system/OS name
from subprocess import call   as system_call  # Execute a shell command

#clases auxiliares
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

#funciones auxiliares
def toJSON(objecttr,value):
	if "{" in value:
		par_json="{\""+objecttr+"\":"+value+"}"
		return par_json
	else:
		par_json="{\""+objecttr+"\":\""+value+"\"}"
		return par_json

def ping(host):
	# Ping command count option as function of OS
	param = '-n' if system_name().lower()=='windows' else '-c'
	# Building the command. Ex: "ping -c 1 google.com"
	command = ['ping', param, '1', host]
	# Pinging
	return system_call(command) == 0

def getQuery(conn,url):
	response=StringIO()
	conn.setopt(conn.URL,url)
	conn.setopt(conn.WRITEFUNCTION,response.write)
	conn.perform()
	conn.close()
	return response

def postQuery(conn,url,data):
	response=StringIO()
	conn.setopt(conn.URL,url)
	conn.setopt(conn.POST, 1)
	conn.setopt(conn.POSTFIELDS, data)
	conn.setopt(conn.WRITEFUNCTION,response.write)
	conn.perform()
	status_code=conn.getinfo(pycurl.HTTP_CODE)
	conn.close()
	return status_code, response

def mapColNum(col):
	col=col.lower()
	letters=list(col)
	if len(letters)>1:
		col_num=((int(ord(letters[0]))-96)*26)+(int(ord(letters[1]))-96)
		return col_num
	else:
		col_num=int(ord(letters[0]))-96
		return col_num

def dictSerialToDevID(db):
	wb=openpyxl.load_workbook(db)
	data=wb['Sheet1']
	dict_serial_to_dev_id={}
	for i in range(2,data.max_row):
		dict_serial_to_dev_id[str(data.cell(i,mapColNum('a')).value)]=str(data.cell(i,mapColNum('b')).value)
	return dict_serial_to_dev_id

def setPostData(action,device,param):
	serial_number=device.split("-")[2]
	if action=="set" and param=='file':
		try:
			config_id=str(DICT_SERIAL_DEV[serial_number])
			with open ('scripts_auxiliares/configs/'+config_id+'.txt') as json_file:
				config=json.load(json_file)
				data_array=[]
				for key in config.keys():
					key_array=[]
					key_array.append(str(key))
					key_array.append(str(config[key]))
					if (str(config[key])=='False') or (str(config[key])=='True'):
						key_array.append('xsd:boolean')
					else:
						key_array.append('xsd:string')
					#print key_array
					data_array.append(key_array)
				#print data_array
				data=json.dumps({"name":"setParameterValues",
					"parameterValues":data_array})
				return data
		except KeyError,e:
			return 101 #codigo de que no hay archivo de config
	else:
		return 102 #codigo de error desconocido

def setRefData(param):
	data=json.dumps({"name":"refreshObject","objectName":str(param)})
	return data

def setPostDict(param):
	data_array=[]
	for key in param.keys():
		key_array=[]
		key_array.append(str(key))
		key_array.append(str(param[key]))
		if param[key]=='InternetGatewayDevice.LANDevice.1.LANHostConfigManagement.DHCPServerEnable':
			key_array.append('xsd:boolean')
		else:
			key_array.append('xsd:string')
		data_array.append(key_array)
	data=json.dumps({"name":"setParameterValues",
	"parameterValues":data_array})	
	return data

#variables globales
global DICT_SERIAL_DEV
global REPORT_PARAMS
DICT_SERIAL_DEV=dictSerialToDevID('serial_dev.xlsx')
REPORT_PARAMS=["InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.URI","InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.URI"]

#vistas
@app.route('/', methods=['GET','POST'])
@app.route('/index', methods=['GET','POST'])
@app.route('/configure', methods=['GET','POST'])
def configure():
	#primera vez que se carga la pagina
    configure_form = ConfigureForm()
    confirm_form=ConfirmForm()
    goback_form=GoBackForm()
    if configure_form.validate_on_submit():
    	#segunda vez que se carga la pagina (haciendo click en configurar)
    	online_devices=[]
    	c=pycurl.Curl()
    	last_inform="_lastInform"
    	now=datetime.utcnow()-timedelta(minutes=200)
    	now=now.isoformat()
    	url_base="http://localhost:7557/devices/"
    	query=toJSON(last_inform,toJSON("$gt",now))
    	query_encoded=urllib.urlencode({"query":query})
    	url=url_base+"?"+query_encoded
    	d=getQuery(c,url).getvalue()
    	json_list=json.loads(d)
    	for j in json_list:
    		online_devices.append(j['_id'])
    	if len(online_devices)<1:
    		return render_template('configure.html', title='Configura', form=configure_form)
    	online_devices=json.dumps(online_devices)
    	return redirect(url_for('confirm_config',online_devices=online_devices))
    return render_template('configure.html', title='Configura', form=configure_form)

@app.route('/verify', methods=['GET','POST'])
def verify():
    form = VerifyForm()
    if form.validate_on_submit():
        url_serial='/verify/'+form.serial.data
        flash(url_serial)
        #sera return redirect(url_serial)
        return redirect(url_for('configure'))
    return render_template('verify.html', title='Verifica', form=form)

@app.route('/confirm_config/<online_devices>', methods=['GET','POST'])
def confirm_config(online_devices):
	confirm_form=ConfirmForm()
	online_devices=json.loads(online_devices)
	if confirm_form.validate_on_submit():
		#aqui se realiza la config de todos los equipos
		url_base="http://localhost:7557/devices/"
		connection_request="/tasks?connection_request"
		devices_ok=[]
		devices_fail=[]
		devices_no_config=[]
		for device in online_devices:
			url=url_base+device+connection_request
			error_flag=0
			c=pycurl.Curl()
			data=setRefData('InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress')
			response=postQuery(c,url,data)
			if response[0]==200:
				try:
					config_id=str(DICT_SERIAL_DEV[device.split("-")[2]])
					try:
						with open('scripts_auxiliares/configs/'+DICT_SERIAL_DEV[device.split("-")[2]]+".txt") as json_file:
							config=json.load(json_file)
							print
							print bcolors.HEADER+"Dispositivo: "+bcolors.ENDC+bcolors.UNDERLINE+device+bcolors.ENDC
							for key in config.keys():
								#convierte cada llave a del dict a un dict
								param_dict={}
								param_dict.update({key:config[key]})
								#revisa que no se cambie el server tr069
								if key!='InternetGatewayDevice.ManagementServer.URL':
									#prepara los datos
									data=setPostDict(param_dict)
									tries=0
									#realiza 5 intentos para la conf de cada param
									while tries<4:
										c=pycurl.Curl()
										response=postQuery(c,url,data)
										#registra respuesta de cada parametro
										if response[0]==200:
											print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.BOLD+" Valor: "+bcolors.ENDC+str(param_dict[key])+bcolors.BOLD+" Respuesta: "+bcolors.ENDC+bcolors.OKGREEN+"OK"+bcolors.ENDC
											break
										elif response[0]==202:
											print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.BOLD+" Valor: "+bcolors.ENDC+str(param_dict[key])+bcolors.BOLD+" Respuesta: "+bcolors.ENDC+bcolors.FAIL+"ERROR"+bcolors.ENDC
											if tries==3:
												print "Este parametro no se pudo configurar."
												tries+=1
												error_flag=1
											elif tries<3:
												print bcolors.WARNING+"Intentando de nuevo ..."+bcolors.ENDC
												c=pycurl.Curl()
												data=setRefData(key)
												postQuery(c,url,data)
												tries+=1
							print "Espera mientras se refrescan los datos configurados."
							#hace un refresh de todos los datos que se configuraron
							for key in config.keys():
								if key!='InternetGatewayDevice.ManagementServer.URL':
									#prepara los datos
									tries=0
									while tries<5:
										c=pycurl.Curl()
										data=setRefData(key)
										response=postQuery(c,url,data)
										if response[0]==200:
											break
										elif response[0]==202:
											#si despues de 5 intentos no se puede refrescar, se manda a fail
											if tries==4:
												print bcolors.WARNING+"Hay un error en la comuniacion con el dispositivo."+bcolors.ENDC
												tries+=1
												error_flag=1
											elif tries<4:
												print "En proceso ..."
												tries+=1
							#se compara cada parametro obtenido del device con el la config aplicada
							for key in config.keys():
								if key=='InternetGatewayDevice.UserInterface.X_HUAWEI_Web.UserInfo.1.Userpassword':
									print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.OKBLUE+" NO ES POSIBLE VERIFICAR CONTRASENAS"+bcolors.ENDC
								elif key=='InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.AuthPassword':
									print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.OKBLUE+" NO ES POSIBLE VERIFICAR CONTRASENAS"+bcolors.ENDC
								elif key=='InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.AuthPassword':
									print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.OKBLUE+" NO ES POSIBLE VERIFICAR CONTRASENAS"+bcolors.ENDC
								elif key!='InternetGatewayDevice.ManagementServer.URL':
									tries=0
									while tries<5:
										c=pycurl.Curl()
										url_base="http://localhost:7557/devices/"
										query=toJSON("_id",device)
										query_encoded=urllib.urlencode({"query":query})
										url=url_base+"?"+query_encoded+"&projection="+key
										param_value=getQuery(c,url).getvalue()
										j=json.loads(param_value)[0]
										param_array=key.split('.')
										for each in param_array:
											j=j[each]
										if j['_value']==config[key]:
											print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.OKGREEN+" VERIFICADO"+bcolors.ENDC
											break
										else:
											if tries==4:
												print bcolors.BOLD+"Parametro: "+bcolors.ENDC+key+bcolors.WARNING+" NO VERIFICADO"+bcolors.ENDC
												error_flag=1
												tries+=1
											elif tries<4:
												time.sleep(2)
												c=pycurl.Curl()
												url=url_base+device+connection_request
												data=setRefData(key)
												postQuery(c,url,data)
												tries+=1
							if error_flag==0:
								print 'Dispositivo '+bcolors.UNDERLINE+device+bcolors.ENDC+' configurado correctamente.'
								print bcolors.WARNING+'Se configurara servidor de gestion, con lo que se pierde la conectividad con el equipo.'+bcolors.ENDC
								tries=0
								while tries<5:
									c=pycurl.Curl()
									data=setPostDict({'InternetGatewayDevice.ManagementServer.URL':'https://172.20.111.202'})
									url=url_base+device+connection_request
									response=postQuery(c,url,data)
									#registra respuesta de cada parametro
									if response[0]==200:
										print bcolors.BOLD+"Parametro: "+bcolors.ENDC+'InternetGatewayDevice.ManagementServer.URL'+bcolors.BOLD+" Valor: "+bcolors.ENDC+'https://172.20.111.202'+bcolors.BOLD+" Respuesta: "+bcolors.ENDC+bcolors.OKGREEN+"OK"+bcolors.ENDC
										print 'Dispositivo '+bcolors.UNDERLINE+device+bcolors.ENDC+' configurado con '+bcolors.OKGREEN+'EXITO'+bcolors.ENDC+'.'
										devices_ok.append(device)
										break
									elif response[0]==202:
										print bcolors.BOLD+"Parametro: "+bcolors.ENDC+'InternetGatewayDevice.ManagementServer.URL'+bcolors.BOLD+" Valor: "+bcolors.ENDC+'https://172.20.111.202'+bcolors.BOLD+" Respuesta: "+bcolors.ENDC+bcolors.FAIL+"ERROR"+bcolors.ENDC
										if tries==4:
											print 'No fue posible configurar el servidor de gestion.'+bcolors.FAIL+' ERROR'+bcolors.ENDC
											tries+=1
											devices_fail.append(device)
										elif tries<4:
											print bcolors.WARNING+"Intentando de nuevo ..."+bcolors.ENDC
											tries+=1
							elif error_flag==1:
								print bcolors.UNDERLINE+device+bcolors.ENDC+' presento errores en la configuracion o verificacion.'
								print 'No se realizara configuracion de servidor de gestion.'
								print bcolors.WARNING+'Se sugiere configurar de nuevo.'+bcolors.ENDC
								devices_fail.append(device)
					except IOError,e:
						#quiere decir que el archivo no existe
						devices_no_config.append(device)
						print
						print bcolors.WARNING+'No existe archivo de configuracion para este equipo: '+bcolors.ENDC+bcolors.UNDERLINE+device+bcolors.ENDC
						print bcolors.FAIL+'ERROR '+bcolors.ENDC+'Equipo no configurado.'
				except KeyError,e:
					#quiere decir que el device id no se encuentra en el excel serial_dev.xlsx
					print
					print bcolors.WARNING+'Este equipo no tiene ninguna configuracion asignada: '+bcolors.ENDC+bcolors.UNDERLINE+device+bcolors.ENDC
					print bcolors.FAIL+'ERROR '+bcolors.ENDC+'Equipo no configurado.'
					devices_no_config.append(device)
				#print error_flag
			else:
				print
				print bcolors.WARNING+'No hay conectividad con el equipo: '+bcolors.ENDC+bcolors.UNDERLINE+device+bcolors.ENDC
				print bcolors.FAIL+'ERROR '+bcolors.ENDC+'Equipo no configurado.'
				devices_fail.append(device)
		devices_ok=json.dumps(devices_ok)
		devices_fail=json.dumps(devices_fail)
		devices_no_config=json.dumps(devices_no_config)
		return redirect(url_for('reporte',devices_ok=devices_ok,devices_fail=devices_fail,devices_no_config=devices_no_config))
	return render_template('confirm_config.html', title='Confirma', devices=online_devices, form=confirm_form)

@app.route('/reporte/<devices_ok>/<devices_fail>/<devices_no_config>',methods=['GET','POST'])
def reporte(devices_ok,devices_fail,devices_no_config):
	goback_form=GoBackForm()
	devices_ok=json.loads(devices_ok)
	devices_fail=json.loads(devices_fail)
	devices_no_config=json.loads(devices_no_config)
	url_base="http://localhost:7557/devices/"
	connection_request="/tasks?connection_request"
	if goback_form.validate_on_submit():
		return redirect(url_for('configure'))
	return render_template('reporte.html',devices_ok_config=devices_ok,devices_fail=devices_fail,devices_no_config=devices_no_config,form=goback_form)
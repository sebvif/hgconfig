import json
import pycurl
from StringIO import StringIO
import urllib
from datetime import datetime,timedelta
import xlrd
import openpyxl
import logging

def toJSON(objecttr,value):
	if "{" in value:
		par_json="{\""+objecttr+"\":"+value+"}"
		return par_json
	else:
		par_json="{\""+objecttr+"\":\""+value+"\"}"
		return par_json

def getQuery(conn,url,query_encoded):
	response=StringIO()
	conn.setopt(conn.URL,url)
	conn.setopt(conn.WRITEDATA,response)
	conn.perform()
	conn.close()
	return response

def postQuery(conn,url,data):
	response=StringIO()
	conn.setopt(conn.URL,url)
	conn.setopt(conn.WRITEFUNCTION,response.write)
	conn.setopt(conn.POST, 1)
	conn.setopt(conn.POSTFIELDS, data)
	conn.perform()
	print response.getvalue()
	print conn.getinfo(pycurl.HTTP_CODE)
	conn.close()
	return response

def recalculateIPPlusOne(ip):
    ups_tarjeta = ''
    for i in range(0, len(ip.split('.'))):
        if i < len(ip.split('.')) - 1:
            ups_tarjeta = ups_tarjeta + ip.split('.')[i]
            ups_tarjeta = ups_tarjeta + '.'
        else:
            try:
                ups_tarjeta = ups_tarjeta + str(int(ip.split('.')[i])+1)
            except ValueError:
                return ip
    return ups_tarjeta

def mapColNum(col):
	col=col.lower()
	letters=list(col)
	if len(letters)>1:
		col_num=((int(ord(letters[0]))-96)*26)+(int(ord(letters[1]))-96)
		return col_num
	else:
		col_num=int(ord(letters[0]))-96
		return col_num

def data_agencias(db):
	wb=openpyxl.load_workbook(db)
	data=wb['Hoja1']
	row_label_list=[]
	config_data_list={}
	last_saved_id=''
	for i in range(4,data.max_row):
		if isinstance(data.cell(i,2).value,unicode) and str(data.cell(i,14).value).lower()=='agencia':
			current_id=data.cell(i,2).value.encode('ascii')
			row_label_list.append(current_id)
			# if data.cell(i,8).value and data.cell(i,9).value:
			# 	dato={'user':data.cell(i,8).value.encode('ascii'),'pass':data.cell(i,9).value.encode('ascii')}
			# else:
			# 	dato={'user':'','pass':''}
			# 	getDBDataLogger.warning('En esta fila no hay usuario o password: '+str(i))
			if current_id!=last_saved_id:
				config_data_list[current_id]={
				#'lines':[dato],
				'cve_inegi':str(data.cell(i,mapColNum('c')).value),
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.AuthUserName':'+52'+str(data.cell(i,mapColNum('h')).value)+'@ims.telmex.com',
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.AuthPassword':str(data.cell(i,mapColNum('i')).value),
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.SIP.URI':'+52'+str(data.cell(i,mapColNum('k')).value),
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.Enable':'Enabled',
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.1.X_HUAWEI_DTMFMethod':'InBand',
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.AuthUserName':'+52'+str(data.cell(i+1,mapColNum('h')).value)+'@ims.telmex.com',
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.AuthPassword':str(data.cell(i+1,mapColNum('i')).value),
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.SIP.URI':'+52'+str(data.cell(i+1,mapColNum('k')).value),
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.Enable':'Enabled',
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.Line.2.X_HUAWEI_DTMFMethod':'InBand',
				# 'InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress':str(data.cell(i,mapColNum('bl')).value),
				# 'InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.DNSServers':str(data.cell(i,mapColNum('an')).value),
				# 'InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.DefaultGateway':str(data.cell(i,mapColNum('bk')).value),
				# 'InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.SubnetMask':'255.255.255.240',
				'InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.AddressingType':'DHCP',
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.SIP.OutboundProxy':str(data.cell(i,mapColNum('ap')).value),
				'InternetGatewayDevice.Services.VoiceService.1.VoiceProfile.1.DigitMap':'*4001*' if str(data.cell(i,mapColNum('bv')).value)=='PREPAGO' else '[1-9]XXXXXXX|***XX|XX*X.#|01XXXXXXXXXX|04[4-5]XXXXXXXXXX|001XXXXXXXXXX|*86|*88|02[1-9]XXXXXXXXX|020|030|03[2-9]|031[0-2]XXX|031[3-4]XXXX|040|0[5-8]X|090|091XXXXXXXXXX|09[2-9]XX.T|00[2-9]XX.T|*#XX*XX.*XX.#|*#XX*XX.#|*#XX#|*#56*|*#26*|*[0-79]X#|*[0-79]X*XX.#|*[0-79]X*XX.*XX.#|*[0-79]X*XX.*XX.*XX.#|*56*|*66*|*69|*89|#XX#|#XX*XX.#|#XX*XX.*XX.#|#*XX*XX.*XX.#|#*XX*XX.#|**X#|163|96163|1471|1475|14713|*34*1234|[1-9]X|[1-9]XX|[1-9]XXX|[1-9]XXXX|[1-9]XXXXX|[1-9]XXXXXX|*222XXX|*9999|**XX|#X.T|##X.T|**X.T|*[0-79]X.T|*8[0-579]X.T|*8[0-579].T',
				'InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.PreSharedKey.1.PreSharedKey ':'m4g4n3m3nt',
				'InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.SSID':'Elara_mgmt',
				'InternetGatewayDevice.LANDevice.1.LANHostConfigManagement.DHCPServerEnable':False,
				'InternetGatewayDevice.UserInterface.X_HUAWEI_Web.UserInfo.1.Username':'ELARA',
				'InternetGatewayDevice.UserInterface.X_HUAWEI_Web.UserInfo.1.Userpassword':'El4r4com!'}
#			else:
#				i=i+1
#				print i
			last_saved_id=current_id
		else:
			pass
	return config_data_list

# c=pycurl.Curl()

# tr_object="_lastInform"
# now=datetime.utcnow()-timedelta(minutes=15)
# now=now.isoformat()

# url_base="http://localhost:7557/devices/"
# query=toJSON(tr_object,toJSON("$lt",now))

# print query

# query_encoded=urllib.urlencode({"query":query})
# url=url_base+"?"+query_encoded

# print url

# d=getQuery(c,url,query_encoded).getvalue()
# j=json.loads(d)
# print type(j)

# for json in j:
# 	print type(json)
# 	print json['InternetGatewayDevice']

#print json.dumps(j, indent=4, sort_keys=False)

# url_base="http://localhost:7557/devices/"
# connection_reques		t="/tasks?connection_request"
# configured_devices=[]
# device="00E0FC-HG659-HW1818L9K2VJ"
# c=pycurl.Curl()
# data=json.dumps({"name":"setParameterValues",
# 				"parameterValues":[
# 					["InternetGatewayDevice.WANDevice.3.WANConnectionDevice.1.WANIPConnection.1.SubnetMask",
# 					"255.255.255.0",
# 					"xsd:string"]
# 				]})
# #ata=json.dumps({"name":"setParameterValues","parameterValues":[["InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.SSID","PruebaWifi","xsd:string"]]})
# url=url_base+device+connection_request
# d=postQuery(c,url,data).getvalue()
# print d
  #query={
  #   "name":"setParameterValues",
  #   "parameterValues":
  #      [
  #         ["InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.SSID", "GenieACS", "xsd:string"],
  #         ["InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.PreSharedKey.1.PreSharedKey", "hello world", "xsd:string"]
  #      ]
  # }

getDBDataLogger=logging.getLogger('getDBData')
data_agencias=data_agencias('BDPSMTRural_2May18.xlsx')
for site_id in data_agencias:
	with open('configs/'+site_id+'.txt','w') as outfile:
		json.dump(data_agencias[site_id],outfile)
	print site_id+': archivo creado'

#with open('agencias_data.txt','w') as outfile:
#	json.dump(data_agencias,outfile)
#print "listo"
#print type(data_agencias)
#print json.dumps(data_agencias,indent=4,sort_keys=True)

#print mapColNum("AA")

#print mapColNum("F")

#print mapColNum("BX")

#7745961205
#TOnManv0rh
#voipnvcompnx1.telmex.net
#*4001*

